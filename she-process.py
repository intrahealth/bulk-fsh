#!/usr/bin/env python
# coding: utf-8

# In[1]:


# pip install pandas openpyxl jinja2 faker
import pandas as pd
import pathlib
from jinja2 import Template
from faker import Faker
import sys
from datetime import date
import random
from openpyxl import load_workbook

# In[2]:


df_names = pd.read_csv('out.csv')
# print(df_names[:4])
names = df_names['en'].sample(n=1, random_state=1)
# print(names.values)
# print(names.values[0])
# build possible vaccines from data dictionary
ws_vaccines = load_workbook('IMZ-DAK-Data-Dictionary.xlsx')['IMZ.A1 Vaccine Library']
vaccine_codes = dict() # { (vaccine name): list[snomed ct code] }
for row in ws_vaccines.iter_rows(2): # skip header row
    if row[15].value is None: # no SNOMEDCT code
        continue
    if row[2].value not in vaccine_codes:
        vaccine_codes[row[2].value] = [{ "code": row[15].value, "series_name": row[16].value }]
    else:
        vaccine_codes[row[2].value].append({ "code": row[15].value, "series_name": row[16].value })
unique_vaccine_names = len(vaccine_codes.keys())

# In[3]:


# goal: nested lists of form: [index, suffix, birthDate, centre...]
# for now, suffix = name and suffix
fake = Faker()
def genfsh(lang):
    # Fix patients to generate
    generated_patients = [{
        "type": "infant-m",
        "birth_date": fake.date_between(start_date='-24m', end_date='-12m'),
        "adm_gender": 'male',
    }, {
        "type": "infant-f",
        "birth_date": fake.date_between(start_date='-24m', end_date='-12m'),
        "adm_gender": 'female',
    }, {
        "type": "teen-m",
        "birth_date": fake.date_between(start_date='-16y', end_date='-13y'),
        "adm_gender": 'male',
    }, {
        "type": "teen-f",
        "birth_date": fake.date_between(start_date='-16y', end_date='-13y'),
        "adm_gender": 'female',
    }, {
        "type": "adult-m",
        "birth_date": fake.date_between(start_date='-30y', end_date='-25y'),
        "adm_gender": 'male',
    }, {
        "type": "adult-f",
        "birth_date": fake.date_between(start_date='-30y', end_date='-25y'),
        "adm_gender": 'female',
    }]

    # Fix a specific code per type of vaccine
    generated_vaccines = []
    for vaccine_name, codes in vaccine_codes.items():
        random_code = random.choice(codes) # for each vaccine, get random vaccination code
        generated_vaccines.append({
            "vaccineName": vaccine_name.replace(" ", "-"), # replacing spaces so that we can insert this in a FHIR Resource's id
            "vaccineCodeName": random_code["code"],
            "vaccineSeriesName": random_code["series_name"],
            "doseNumberPositiveInt": "TODO: DOSE", # TODO
            "seriesDosesPositiveInt": "TODO: SERIES", # TODO
        })

    # Generate patients
    for patient in generated_patients:
        tempname = df_names[lang].sample(n=1)
        tempname = tempname.values[0]
        suffix = lang + str(patient["type"])
        name = tempname + str(patient["type"])
        if lang == 'en':
            orgname = "Government Hospital"
            centre = 'Vaccination Site'
        if lang == 'es':
            orgname = "Hospital del Gobierno"
            centre = "Sitio de vacunación"
        if lang == 'fr':
            orgname = "Hôpital du gouvernement"
            centre = "Site de vaccination"
        if lang == 'ar':
            orgname = "مستشفى حكومي"
            centre = "موقع التطعيم"
        if lang == 'zh':
            orgname = "政府医院"
            centre = "疫苗接种现场"
        if lang == 'ru':
            orgname = "Государственная больница"
            centre = "Сайт вакцинации"
        birthDate = str(patient['birth_date'])
        date_today = date.today()
        vaccine_date = str(fake.date_between_dates(date_start=patient["birth_date"], date_end=date_today))
        identifier = lang + str(9999) + str(patient["type"])
        adm_gender = patient['adm_gender']

        # # # assign random observations: 14 months, 15, 30 -> each vaccination once (random code)
        # # pregnancy
        # pregnant = False
        # if (adm_gender == "female" and random.random() < 0.5):
        #     pregnant = True
        # # HIV positive
        # hiv_positive = False
        # if (random.random() < 0.4):
        #     hiv_positive = True
        
        # this prints oddly bc of the mix of rtl-ltr langs?
        print(lang, suffix, name, birthDate, identifier, adm_gender)
        # put through jinja2
        path = pathlib.Path('she-template.fsh')
        text = path.read_text()
        t = Template(text)
        msg = t.render(
            suffix=suffix,
            name=name,
            birthDate=birthDate,
            identifier=identifier,
            adm_gender=adm_gender,
            orgname=orgname,
            centre=centre,
            date=vaccine_date,
            date_today=date_today,
            immunizations=generated_vaccines,
            pregnant=False,
            hiv_positive=False,
            gen_extra_data=False
        )
        path_out = pathlib.Path(f"input/fsh/{suffix}.fsh")
        path_out.write_text(msg)


# In[4]:


print("command:", str(sys.argv))
langs = ['en', 'es', 'fr', 'ar', 'zh', 'ru']
if len(sys.argv) > 1 and sys.argv[1] in langs:
    genfsh(str(sys.argv[1]))
else:
    genfsh('ar')


# In[ ]:




